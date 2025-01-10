from pathlib import Path
from openpyxl import load_workbook
import csv

data_dir = Path(__file__).parent.parent / "excel_data" # 数据存放目录


def read_csv(filename):
    path = data_dir / filename

    with open(path, encoding="utf-8") as f:
        reader = csv.DictReader(f)

        for i in reader:
            # print(i)
            yield list(i.values())  # 小重点, yeild 起到返回值的效果
            """yeild自己理解的就是不太硬性的return返回值，返回后还会在去执行该函数的下一步代码"""


def read_excel(filename):
    path = data_dir / filename
    print(path)
    # wb = load_workbook(path)
    # ws = wb.active        #  excel的页表sheet1

    # yield from ws.iter_rows(min_row=2,values_only=True)             # 从第2行开始  # 只要excel中的值
    if not path.exists():
        print("File does not exist.")
        return
    try:
        wb = load_workbook(path)
        ws = wb.active  # 获取活动工作表
        for row in ws.iter_rows(min_row=2, values_only=True):
            yield row
    except Exception as e:
        print(f"An error occurred: {e}")


if __name__ == "__main__":
    for d in read_excel("ddt-test_login.xlsx"):
        print(d)